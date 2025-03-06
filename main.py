import pandas as pd
import os
from threading import Thread
from openpyxl import load_workbook

# Campos requeridos
CAMPOS = {
    "Fecha de inicio": "Inicio_de_Conexión_Dia",
    "Usuario": "Usuario"
}

# Verifica si una fecha es día no laboral
def es_dia_no_laboral(fecha, feriados):
    if pd.isna(fecha):  # Fecha vacía no es día no laboral
        return False
    return fecha.weekday() in [5, 6] or fecha in feriados

# Filtra por rango de fechas
def filtrar_por_rango_fechas(datos, fecha_inicio, fecha_fin):
    return datos[(datos['Inicio_de_Conexión_Dia'] >= fecha_inicio) & (datos['Inicio_de_Conexión_Dia'] <= fecha_fin)]

# Optimiza el DataFrame convirtiendo columnas a tipos ligeros
def optimizar_dataframe(datos):
    datos['Usuario'] = datos['Usuario'].astype('category')
    return datos

# Lee datos por partes usando openpyxl
def leer_datos_por_partes(ruta_archivo, usecols, chunk_size=50000):
    wb = load_workbook(ruta_archivo, read_only=True)
    ws = wb.active
    encabezados = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    columnas_indices = [encabezados.index(col) for col in usecols]

    datos_chunks = []
    chunk = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        chunk.append([row[idx] for idx in columnas_indices])
        if len(chunk) == chunk_size:
            datos_chunks.append(pd.DataFrame(chunk, columns=usecols))
            chunk = []
    if chunk:
        datos_chunks.append(pd.DataFrame(chunk, columns=usecols))

    wb.close()
    return datos_chunks

# Procesa datos con flujo ajustado para vista previa y confirmación
def procesar_datos(ruta_archivo, feriados, fecha_inicio=None, fecha_fin=None):
    try:
        print("Cargando datos...")
        # Leer datos en chunks
        chunks = leer_datos_por_partes(ruta_archivo, list(CAMPOS.values()))

        datos_filtrados = []

        for chunk in chunks:
            chunk['Inicio_de_Conexión_Dia'] = pd.to_datetime(chunk['Inicio_de_Conexión_Dia'], errors='coerce')

            if fecha_inicio and fecha_fin:
                chunk = filtrar_por_rango_fechas(chunk, fecha_inicio, fecha_fin)

            chunk = chunk[chunk['Inicio_de_Conexión_Dia'].apply(lambda x: es_dia_no_laboral(x, feriados))]
            datos_filtrados.append(chunk)

        # Concatenar todos los chunks filtrados
        datos_filtrados = pd.concat(datos_filtrados, ignore_index=True)
        datos_filtrados = optimizar_dataframe(datos_filtrados)

        # Mostrar vista previa antes de exportar
        print("\nVista previa de los datos filtrados:")
        print(datos_filtrados.head(10))

        # Confirmar exportación
        confirmar = input("\u00bfDeseas exportar los datos a Excel? (s/n): ").strip().lower()
        if confirmar == 's':
            archivo_salida = 'conexiones_feriados_y_fines_de_semana.xlsx'
            datos_filtrados.to_excel(archivo_salida, index=False)
            print(f"Resultados exportados a {archivo_salida}")
        else:
            print("Exportación cancelada por el usuario.")

    except Exception as e:
        print(f"Error al procesar los datos: {e}")

# Procesamiento en segundo plano
def procesar_en_segundo_plano(ruta_archivo, feriados, fecha_inicio, fecha_fin):
    hilo = Thread(target=procesar_datos, args=(ruta_archivo, feriados, fecha_inicio, fecha_fin))
    hilo.start()
    print("Procesando en segundo plano...")

# Función principal
def main():
    # Ruta del archivo en la carpeta del proyecto
    ruta_archivo = os.path.join(os.getcwd(), "bd_automatas.xlsx")

    if not os.path.exists(ruta_archivo):
        print("Error: El archivo 'bd_automatas.xlsx' no se encuentra en la carpeta del proyecto.")
        return

    # Lista de feriados (ejemplo)
    feriados = pd.to_datetime([
        '2019-01-01', '2019-02-24', '2019-03-29', '2019-05-01',
        '2019-06-20', '2019-07-09', '2019-08-17', '2019-10-12',
        '2019-11-02', '2019-12-25'
    ])

    # Solicitar rango de fechas al usuario
    fecha_inicio_str = input("Ingresa la fecha de inicio (YYYY-MM-DD, presiona Enter para omitir): ")
    fecha_fin_str = input("Ingresa la fecha de fin (YYYY-MM-DD, presiona Enter para omitir): ")
    
    fecha_inicio = pd.to_datetime(fecha_inicio_str) if fecha_inicio_str else None
    fecha_fin = pd.to_datetime(fecha_fin_str) if fecha_fin_str else None

    # Validar el rango de fechas
    if fecha_inicio and fecha_fin and fecha_inicio > fecha_fin:
        print("Error: La fecha de inicio no puede ser mayor a la fecha de fin.")
        return

    # Procesar los datos
    procesar_en_segundo_plano(ruta_archivo, feriados, fecha_inicio, fecha_fin)

if __name__ == "__main__":
    main()
